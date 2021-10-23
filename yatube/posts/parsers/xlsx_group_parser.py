from openpyxl import load_workbook


class Group:
    def __init__(self, groups_data):
        self.title = groups_data['title']
        self.slug = groups_data['slug']
        self.description = groups_data['description']

    def __str__(self):
        return self.title


def xlsx_group_parser(file):
    wb = load_workbook(filename=file, read_only=True)
    ws = wb.active
    groups_data = [Group(row_parser(row)) for row in ws.iter_rows(min_row=2)]
    wb.close()
    return groups_data


def row_parser(row):
    groups_data = dict.fromkeys(['title', 'slug', 'description'])
    count = 0
    for data in groups_data:
        groups_data[data] = row[count].value
        count += 1
    return groups_data


def here():
    print('here')

# if __name__ == '__main__':
#     wb = load_workbook(filename='groups.xlsx', read_only=True)
#     group = XLSXGroupParser(wb)
